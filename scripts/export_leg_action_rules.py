#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Экспорт правил вилки из «Действия статусы.xlsx» в config/order_status/*.json.

Источник истины — Excel. Ручное редактирование JSON не нужно:
  python scripts/export_leg_action_rules.py "C:/path/Действия статусы.xlsx"

Без аргумента — читает tests/fixtures/order_status_Безнал.csv и order_status_Нал.csv
(их тоже можно обновить из xlsx флагом --sync-csv).
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

ACTIONS = {"опрос", "отмена", "востановление", "ничего"}
STATES = {
    "SearchesForCar",
    "WaitingCarSearch",
    "CarFound",
    "Running",
    "Canceled",
    "Executed",
    "CostCalculation",
}

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "tests" / "fixtures"
OUT = ROOT / "config" / "order_status"
DEFAULT_XLSX = Path(r"c:\Users\user\Documents\MEGA\Documents\Такси\Действия статусы.xlsx")

SHEET_BEZNAL = "Безнал"
SHEET_NAL = "Нал"
SHEET_MATRIX = "Лист1"


def split_glued_token(token: str):
    remaining = token
    ordered = sorted(STATES, key=len, reverse=True)
    result = []
    while remaining:
        matched = False
        for state in ordered:
            if remaining.startswith(state):
                result.append(state)
                remaining = remaining[len(state) :]
                matched = True
                break
        if not matched:
            break
    return result


def split_states(text: str):
    if text is None:
        return []
    text = str(text).strip()
    if not text or text in {"любой", "любая", "*"}:
        return []
    result = []
    for token in re.split(r"[\s,]+", text):
        token = token.strip()
        if not token:
            continue
        if token in STATES:
            result.append(token)
        else:
            result.extend(split_glued_token(token))
    deduped = []
    for state in result:
        if state not in deduped:
            deduped.append(state)
    return deduped


def find_actions(row):
    found = []
    for cell in row:
        if cell is None:
            continue
        value = str(cell).strip()
        if value in ACTIONS:
            found.append(value)
    if len(found) >= 2:
        return found[0], found[1]
    if len(found) == 1:
        return found[0], "ничего"
    return None, None


def parse_beznal_rows(rows):
    rules = []
    index_by_key = {}
    cur_bonus = None
    cur_double_opts = None
    cur_last_bonus_opts = None

    for row in rows[2:]:
        if not row or all(not str(c).strip() if c is not None else True for c in row):
            cur_bonus = None
            cur_double_opts = None
            cur_last_bonus_opts = None
            continue

        bonus_action, double_action = find_actions(row)
        if not bonus_action:
            continue

        col0 = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        col1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        col2 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        if col0 in STATES:
            cur_bonus = col0
        if col1 and split_states(col1):
            cur_double_opts = split_states(col1)
        if col2:
            cur_last_bonus_opts = split_states(col2) or None

        if not cur_bonus:
            continue

        item = {
            "bonus": cur_bonus,
            "double_options": cur_double_opts or ["*"],
            "bonus_action": bonus_action,
            "double_action": double_action,
        }
        if cur_last_bonus_opts:
            item["last_bonus_options"] = cur_last_bonus_opts

        key = (
            item["bonus"],
            tuple(item["double_options"]),
            tuple(item.get("last_bonus_options", ("*",))),
        )
        if key in index_by_key:
            rules[index_by_key[key]] = item
        else:
            index_by_key[key] = len(rules)
            rules.append(item)

    return rules


def parse_nal_rows(rows):
    rules = []
    index_by_key = {}
    cur_double = None
    cur_bonus_opts = None
    cur_last_double_opts = None

    for row in rows[2:]:
        if not row or all(not str(c).strip() if c is not None else True for c in row):
            cur_double = None
            cur_bonus_opts = None
            cur_last_double_opts = None
            continue

        double_action, bonus_action = find_actions(row)
        if not double_action:
            continue

        col0 = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        col1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        col2 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        if col0 in STATES:
            cur_double = col0
        if col1 and split_states(col1):
            cur_bonus_opts = split_states(col1)
        if col2:
            cur_last_double_opts = split_states(col2) or None

        if not cur_double:
            continue

        item = {
            "double": cur_double,
            "bonus_options": cur_bonus_opts or ["*"],
            "double_action": double_action,
            "bonus_action": bonus_action,
        }
        if cur_last_double_opts:
            item["last_double_options"] = cur_last_double_opts

        key = (
            item["double"],
            tuple(item["bonus_options"]),
            tuple(item.get("last_double_options", ("*",))),
        )
        if key in index_by_key:
            rules[index_by_key[key]] = item
        else:
            index_by_key[key] = len(rules)
            rules.append(item)

    return rules


def sheet_to_rows(ws):
    return [list(row) for row in ws.iter_rows(values_only=True)]


def read_xlsx(path: Path):
    if load_workbook is None:
        raise RuntimeError("openpyxl не установлен: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in (SHEET_BEZNAL, SHEET_NAL, SHEET_MATRIX):
        if name in wb.sheetnames:
            sheets[name] = sheet_to_rows(wb[name])
    wb.close()
    return sheets


def write_csv(path: Path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        for row in rows:
            writer.writerow(["" if c is None else c for c in row])


def cases_from_bonus_rules(rules):
    cases = []
    seen = set()
    for rule in rules:
        doubles = rule["double_options"]
        if "*" in doubles:
            doubles = sorted(STATES)
        lasts = rule.get("last_bonus_options") or [None]
        for double_state in doubles:
            for last in lasts:
                case = {
                    "bonus": rule["bonus"],
                    "double": double_state,
                    "last_bonus": last,
                    "bonus_action": rule["bonus_action"],
                    "double_action": rule["double_action"],
                }
                key = tuple(case.values())
                if key in seen:
                    continue
                seen.add(key)
                cases.append(case)
    return cases


def cases_from_nal_rules(rules):
    cases = []
    seen = set()
    for rule in rules:
        bonuses = rule["bonus_options"]
        if "*" in bonuses:
            bonuses = sorted(STATES)
        lasts = rule.get("last_double_options") or [None]
        for bonus_state in bonuses:
            for last in lasts:
                case = {
                    "double": rule["double"],
                    "bonus": bonus_state,
                    "last_double": last,
                    "double_action": rule["double_action"],
                    "bonus_action": rule["bonus_action"],
                }
                key = tuple(case.values())
                if key in seen:
                    continue
                seen.add(key)
                cases.append(case)
    return cases


def parse_beznal_cases(rows):
    return cases_from_bonus_rules(parse_beznal_rows(rows))


def parse_nal_cases(rows):
    return cases_from_nal_rules(parse_nal_rows(rows))


def parse_matrix_sheet(rows):
    """Лист1 — полная матрица для автотестов (каждая строка = один сценарий)."""
    cases = []
    seen = set()

    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        cells = [str(c).strip() if c is not None else "" for c in row]
        double_action, bonus_action = find_actions(cells)
        if not double_action:
            continue

        nal_state = ""
        bonus_raw = ""
        last_nal_raw = ""
        last_bonus_raw = ""

        i = 0
        while i < len(cells):
            label = cells[i]
            value = cells[i + 1] if i + 1 < len(cells) else ""
            if label == "текущее нал" and value:
                states = split_states(value)
                if states:
                    nal_state = states[0]
            elif label == "текущее безнал" and value:
                bonus_raw = value
            elif label == "предыд нал" and value:
                last_nal_raw = value
            elif label == "предыд безнал" and value:
                last_bonus_raw = value
            i += 1

        bonus_states = split_states(bonus_raw)
        if not nal_state or not bonus_states:
            continue

        last_nal_states = split_states(last_nal_raw) or [None]
        if last_bonus_raw in {"", "любой", "любая", "*"}:
            last_bonus_states = [None]
        else:
            last_bonus_states = split_states(last_bonus_raw) or [None]

        for bonus_state in bonus_states:
            for last_nal in last_nal_states:
                for last_bonus in last_bonus_states:
                    case = {
                        "nal": nal_state,
                        "bonus": bonus_state,
                        "last_nal": last_nal,
                        "last_bonus": last_bonus,
                        "nal_action": double_action,
                        "bonus_action": bonus_action,
                    }
                    key = tuple(case.values())
                    if key in seen:
                        continue
                    seen.add(key)
                    cases.append(case)

    return cases


def export_all(xlsx_path: Path | None, sync_csv: bool):
    if xlsx_path and xlsx_path.is_file():
        sheets = read_xlsx(xlsx_path)
        bez_rows = sheets[SHEET_BEZNAL]
        nal_rows = sheets[SHEET_NAL]
        matrix_rows = sheets.get(SHEET_MATRIX, [])
        if sync_csv:
            write_csv(FIXTURES / "order_status_Безнал.csv", bez_rows)
            write_csv(FIXTURES / "order_status_Нал.csv", nal_rows)
            if matrix_rows:
                write_csv(FIXTURES / "order_status_Лист1.csv", matrix_rows)
    else:
        bez_rows = list(csv.reader(open(FIXTURES / "order_status_Безнал.csv", encoding="utf-8-sig")))
        nal_rows = list(csv.reader(open(FIXTURES / "order_status_Нал.csv", encoding="utf-8-sig")))
        matrix_rows = list(csv.reader(open(FIXTURES / "order_status_Лист1.csv", encoding="utf-8-sig")))

    bonus_rules = parse_beznal_rows(bez_rows)
    nal_rules = parse_nal_rows(nal_rows)
    bonus_cases = parse_beznal_cases(bez_rows)
    nal_cases = parse_nal_cases(nal_rows)
    matrix_cases = parse_matrix_sheet(matrix_rows)

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "leg_action_bonus.json").write_text(
        json.dumps(bonus_rules, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (OUT / "leg_action_nal.json").write_text(
        json.dumps(nal_rules, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (FIXTURES / "leg_action_bonus_cases.json").write_text(
        json.dumps(bonus_cases, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (FIXTURES / "leg_action_nal_cases.json").write_text(
        json.dumps(nal_cases, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (FIXTURES / "leg_action_matrix_cases.json").write_text(
        json.dumps(matrix_cases, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    return bonus_rules, nal_rules, bonus_cases, nal_cases, matrix_cases


def main():
    parser = argparse.ArgumentParser(description="Export fork leg rules from Excel")
    parser.add_argument(
        "xlsx",
        nargs="?",
        default=str(DEFAULT_XLSX) if DEFAULT_XLSX.is_file() else None,
        help="Path to «Действия статусы.xlsx»",
    )
    parser.add_argument(
        "--sync-csv",
        action="store_true",
        help="Update tests/fixtures/*.csv from xlsx",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx) if args.xlsx else None
    bonus, nal, bonus_cases, nal_cases, matrix_cases = export_all(xlsx_path, args.sync_csv)

    print(f"bonus rules: {len(bonus)}")
    print(f"nal rules: {len(nal)}")
    print(f"bonus phase cases: {len(bonus_cases)}")
    print(f"nal phase cases: {len(nal_cases)}")
    print(f"matrix cases (Лист1 cross-check): {len(matrix_cases)}")
    if xlsx_path:
        print(f"source: {xlsx_path}")
    else:
        print("source: tests/fixtures CSV")


if __name__ == "__main__":
    main()
